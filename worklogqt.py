import sys
import os
import csv
import threading
import socket
import io
from datetime import datetime
from PySide6.QtWidgets import (QApplication, QMainWindow, QTabWidget, QWidget,
                             QPushButton, QGridLayout, QVBoxLayout, QHBoxLayout,
                             QLabel, QTableWidget, QTableWidgetItem,
                             QDialog, QTextEdit, QGroupBox,
                             QFileDialog, QDateEdit,
                             QInputDialog, QLineEdit, QMessageBox, QCheckBox)
from PySide6.QtCore import Qt, QDate, QThread, Signal, Slot, QSettings, QTimer
from PySide6.QtGui import QIcon, QPixmap, QImage

from flask import Flask, request, render_template_string, redirect, url_for, session
from werkzeug.serving import make_server
import qrcode
from PIL import Image

# 全局锁，用于文件写入安全
file_lock = threading.Lock()

# HTML 模板
LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>工作日志 - 登录</title>
    <style>
        body { font-family: Arial, sans-serif; display: flex; justify-content: center; align-items: center; height: 100vh; margin: 0; background-color: #f0f2f5; }
        .container { background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); width: 90%; max-width: 400px; text-align: center; }
        input { width: 100%; padding: 10px; margin: 10px 0; border: 1px solid #ddd; border-radius: 4px; box-sizing: border-box; }
        button { width: 100%; padding: 10px; background-color: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 16px; }
        button:hover { background-color: #0056b3; }
        .error { color: red; margin-bottom: 10px; }
    </style>
</head>
<body>
    <div class="container">
        <h2>请输入访问密码</h2>
        {% if error %}
        <div class="error">{{ error }}</div>
        {% endif %}
        <form method="post">
            <input type="password" name="password" placeholder="密码" required>
            <button type="submit">登录</button>
        </form>
    </div>
</body>
</html>
"""

HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>工作日志记录</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 0; padding: 20px; background-color: #f0f2f5; }
        .container { max-width: 600px; margin: 0 auto; }
        h2 { text-align: center; color: #333; }
        .grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 10px; }
        .btn { padding: 15px; background-color: white; border: 1px solid #ddd; border-radius: 8px; text-align: center; cursor: pointer; font-size: 14px; color: #333; box-shadow: 0 1px 2px rgba(0,0,0,0.05); transition: all 0.2s; }
        .btn:active { background-color: #e9ecef; transform: scale(0.98); }
        .btn-other { background-color: #f8f9fa; color: #007bff; font-weight: bold; }
        .success-modal { display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background-color: rgba(0,0,0,0.5); z-index: 1000; align-items: center; justify-content: center; }
        .success-content { background-color: white; padding: 25px; border-radius: 10px; width: 80%; max-width: 300px; text-align: center; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }
        .success-icon { color: #28a745; font-size: 48px; margin-bottom: 10px; }
        .success-btn { background-color: #28a745; color: white; border: none; padding: 10px 30px; border-radius: 5px; font-size: 16px; cursor: pointer; margin-top: 15px; }
        .success-btn:hover { background-color: #218838; }
        .modal { display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background-color: rgba(0,0,0,0.5); z-index: 1000; }
        .modal-content { background-color: white; margin: 20% auto; padding: 20px; width: 80%; max-width: 400px; border-radius: 8px; }
        textarea { width: 100%; height: 100px; margin: 10px 0; padding: 8px; box-sizing: border-box; border: 1px solid #ddd; border-radius: 4px; }
        .modal-btns { display: flex; justify-content: flex-end; gap: 10px; }
    </style>
    <script>
        function closeSuccessModal() {
            document.getElementById('successModal').style.display = 'none';
            // 移除URL中的查询参数，防止刷新页面重复提交（简单处理，实际可能需要重定向）
            // distinct from typical PRG, but helpful if just closing a modal
        }

        function openOtherDialog() {
            document.getElementById('otherModal').style.display = 'block';
        }
        
        function closeOtherDialog() {
            document.getElementById('otherModal').style.display = 'none';
        }
    </script>
</head>
<body>
    <div class="container">
        <h2>点击记录工作</h2>
        
        {% if success %}
        <div id="successModal" class="success-modal" style="display: flex;">
            <div class="success-content">
                <div class="success-icon">✓</div>
                <h3>记录成功！</h3>
                <button class="success-btn" onclick="closeSuccessModal()">确定</button>
            </div>
        </div>
        {% endif %}
        
        <div class="grid">
            {% for category in categories %}
            <form method="post" style="display: contents;">
                <input type="hidden" name="category" value="{{ category }}">
                {% if category == '其他' %}
                <div class="btn btn-other" onclick="openOtherDialog()">{{ category }}</div>
                {% else %}
                <button type="submit" class="btn">{{ category }}</button>
                {% endif %}
            </form>
            {% endfor %}
        </div>
    </div>

    <div id="otherModal" class="modal">
        <div class="modal-content">
            <h3>输入工作内容</h3>
            <form method="post">
                <input type="hidden" name="category" value="其他">
                <textarea name="content" placeholder="请输入具体工作内容..." required></textarea>
                <div class="modal-btns">
                    <button type="button" class="btn" onclick="closeOtherDialog()" style="width: auto; padding: 8px 15px;">取消</button>
                    <button type="submit" class="btn" style="width: auto; padding: 8px 15px; background-color: #007bff; color: white;">保存</button>
                </div>
            </form>
        </div>
    </div>
</body>
</html>
"""

class MobileServerThread(QThread):
    server_started = Signal(str) # 发送服务器地址
    server_error = Signal(str)

    def __init__(self, log_file, categories, password, port=5000):
        super().__init__()
        self.log_file = log_file
        self.categories = categories
        self.password = password
        self.port = port
        self.app = Flask(__name__)
        self.app.secret_key = os.urandom(24)
        self.server = None
        
        self.setup_routes()

    def setup_routes(self):
        @self.app.route('/', methods=['GET', 'POST'])
        def index():
            if 'logged_in' not in session:
                return redirect(url_for('login'))
            
            if request.method == 'POST':
                category = request.form.get('category')
                content = request.form.get('content', '')
                if category:
                    self.save_log(category, content)
                    return render_template_string(HTML_TEMPLATE, categories=self.categories, success=True)
            
            return render_template_string(HTML_TEMPLATE, categories=self.categories, success=False)

        @self.app.route('/login', methods=['GET', 'POST'])
        def login():
            error = None
            if request.method == 'POST':
                if request.form['password'] == self.password:
                    session['logged_in'] = True
                    return redirect(url_for('index'))
                else:
                    error = '密码错误'
            return render_template_string(LOGIN_TEMPLATE, error=error)

    def save_log(self, category, content):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with file_lock:
                if not os.path.exists(self.log_file):
                    with open(self.log_file, 'w', newline='', encoding='utf-8-sig') as f:
                        writer = csv.writer(f)
                        writer.writerow(['时间', '工作类别', '工作内容'])
                
                with open(self.log_file, 'a', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow([current_time, category, content])
        except Exception as e:
            print(f"Error saving log: {e}")

    def run(self):
        # 查找可用端口
        while True:
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.bind(('', self.port))
                break
            except OSError:
                self.port += 1
        
        # 获取本机IP
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
        except:
            ip = "127.0.0.1"
            
        url = f"http://{ip}:{self.port}"
        self.server_started.emit(url)
        
        try:
            self.server = make_server('0.0.0.0', self.port, self.app)
            self.server.serve_forever()
        except Exception as e:
            self.server_error.emit(str(e))

    def stop(self):
        if self.server:
            self.server.shutdown()

    def update_password(self, new_password):
        self.password = new_password

class WorkLogRecorder(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("工作日志记录器")
        self.resize(800, 600)

        documents_path = os.path.expanduser('~/Documents/WorkLog')
        if not os.path.exists(documents_path):
            os.makedirs(documents_path)
        self.log_file = os.path.join(documents_path, "worklog.csv")

        self.categories = [
            "电脑硬件维修", "电脑软件维修类", "打印机维护", "网络设备维护",
            "安防设备维护", "服务器维护", "硬件测试", "软件测试",
            "OA后台业务维护", "ERP维护", "PLM维护", "CRM维护", "加密系统维护",
            "SMB维护", "云平台业务维护", "电话系统维护", "投影仪维修调试",
            "音响维修调试", "电路维修调试", "咨询服务", "系统重装", "食堂打卡机", "其他"
        ]
        
        self.server_thread = None
        self.server_url = ""

        # 初始化设置
        self.settings = QSettings("MyCompany", "WorkLogRecorder")

        self.init_ui()
        self.load_data()
        
        # 检查自动启动
        self.check_auto_start()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        self.record_tab = QWidget()
        self.tab_widget.addTab(self.record_tab, "记录工作")
        
        self.stats_tab = QWidget()
        self.tab_widget.addTab(self.stats_tab, "统计报表")
        
        # 新增手机同步标签页
        self.mobile_tab = QWidget()
        self.tab_widget.addTab(self.mobile_tab, "手机同步")
        
        self.init_record_tab()
        self.init_stats_tab()
        self.init_mobile_tab()

    def init_record_tab(self):
        record_layout = QVBoxLayout(self.record_tab)
        
        button_group = QGroupBox("选择工作类别")
        button_layout = QGridLayout()
        button_group.setLayout(button_layout)
        
        row, col = 0, 0
        for category in self.categories:
            button = QPushButton(category)
            button.clicked.connect(lambda checked, cat=category: self.log_work(cat))
            button_layout.addWidget(button, row, col)
            col += 1
            if col > 3:
                col = 0
                row += 1
        
        record_layout.addWidget(button_group)
        
        undo_btn = QPushButton("撤销 (1分钟内)")
        undo_btn.setStyleSheet("background-color: #dc3545; color: white; padding: 10px; font-weight: bold; border-radius: 5px;")
        undo_btn.clicked.connect(self.undo_last_log)
        record_layout.addWidget(undo_btn)
        record_layout.addStretch()

    def init_stats_tab(self):
        stats_layout = QVBoxLayout(self.stats_tab)

        filter_group = QGroupBox("统计筛选")
        filter_layout = QGridLayout()
        filter_group.setLayout(filter_layout)

        filter_layout.addWidget(QLabel("开始日期:"), 0, 0)
        self.stats_start_date = QDateEdit()
        self.stats_start_date.setDisplayFormat("yyyy-MM-dd")
        self.stats_start_date.setDate(QDate.currentDate().addDays(-5))
        filter_layout.addWidget(self.stats_start_date, 0, 1)

        filter_layout.addWidget(QLabel("结束日期:"), 0, 2)
        self.stats_end_date = QDateEdit()
        self.stats_end_date.setDisplayFormat("yyyy-MM-dd")
        self.stats_end_date.setDate(QDate.currentDate())
        filter_layout.addWidget(self.stats_end_date, 0, 3)

        stats_btn = QPushButton("生成统计")
        stats_btn.clicked.connect(self.generate_stats)
        filter_layout.addWidget(stats_btn, 0, 4)

        stats_layout.addWidget(filter_group)

        self.stats_table = QTableWidget()
        self.stats_table.setColumnCount(3)
        self.stats_table.setHorizontalHeaderLabels(["工作类别", "次数", "归档统计"])
        self.stats_table.horizontalHeader().setStretchLastSection(True)
        stats_layout.addWidget(self.stats_table)

        export_stats_group = QGroupBox("导出统计")
        export_stats_layout = QHBoxLayout()
        export_stats_group.setLayout(export_stats_layout)

        export_stats_excel_btn = QPushButton("导出Excel")
        export_stats_excel_btn.clicked.connect(self.export_stats_excel)
        export_stats_layout.addWidget(export_stats_excel_btn)

        stats_layout.addWidget(export_stats_group)

    def init_mobile_tab(self):
        layout = QVBoxLayout(self.mobile_tab)
        
        # 控制区域
        control_group = QGroupBox("服务设置")
        control_layout = QGridLayout()
        control_group.setLayout(control_layout)
        
        self.server_btn = QPushButton("开启手机记录")
        self.server_btn.setCheckable(True)
        self.server_btn.toggled.connect(self.toggle_server)
        control_layout.addWidget(self.server_btn, 0, 0, 1, 2)
        
        control_layout.addWidget(QLabel("访问密码:"), 1, 0)
        self.password_edit = QLineEdit("21232")
        self.password_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_edit.textChanged.connect(self.update_password)
        self.password_edit.textChanged.connect(self.update_password)
        control_layout.addWidget(self.password_edit, 1, 1)

        # 自动启动复选框
        self.auto_start_cb = QCheckBox("启动时自动开启 (延迟60秒)")
        self.auto_start_cb.setChecked(self.settings.value("auto_start_mobile_sync", False, type=bool))
        self.auto_start_cb.stateChanged.connect(self.toggle_auto_start)
        control_layout.addWidget(self.auto_start_cb, 2, 0, 1, 2)
        
        layout.addWidget(control_group)
        
        # 信息显示区域
        self.info_group = QGroupBox("连接信息")
        info_layout = QVBoxLayout()
        self.info_group.setLayout(info_layout)
        self.info_group.setVisible(False)
        
        self.url_label = QLabel()
        self.url_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.url_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #007bff; margin: 10px;")
        info_layout.addWidget(self.url_label)
        
        self.qr_label = QLabel()
        self.qr_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.qr_label.setMinimumSize(200, 200)
        info_layout.addWidget(self.qr_label)
        
        tips_label = QLabel("请确保手机和电脑连接同一WiFi")
        tips_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tips_label.setStyleSheet("color: #666;")
        info_layout.addWidget(tips_label)
        
        layout.addWidget(self.info_group)
        layout.addStretch()

    def toggle_server(self, checked):
        if checked:
            self.server_btn.setText("停止手机记录")
            self.password_edit.setEnabled(False)
            
            # 启动服务器线程
            self.server_thread = MobileServerThread(
                self.log_file, 
                self.categories, 
                self.password_edit.text()
            )
            self.server_thread.server_started.connect(self.on_server_started)
            self.server_thread.server_error.connect(self.on_server_error)
            self.server_thread.start()
        else:
            self.server_btn.setText("开启手机记录")
            self.password_edit.setEnabled(True)
            self.info_group.setVisible(False)
            
            if self.server_thread:
                self.server_thread.stop()
                self.server_thread.wait()
                self.server_thread = None

    def on_server_started(self, url):
        self.server_url = url
        self.url_label.setText(f"访问地址: {url}")
        self.generate_qr_code(url)
        self.info_group.setVisible(True)

    def on_server_error(self, error_msg):
        self.server_btn.setChecked(False)
        self.toggle_server(False)
        CustomMessageBox(self, "启动失败", f"无法启动服务器: {error_msg}").exec()

    def generate_qr_code(self, data):
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        # 转换为QPixmap
        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        qimg = QImage.fromData(buffer.getvalue())
        pixmap = QPixmap.fromImage(qimg)
        
        self.qr_label.setPixmap(pixmap.scaled(200, 200, Qt.AspectRatioMode.KeepAspectRatio))

    def update_password(self, text):
        if self.server_thread:
            self.server_thread.update_password(text)

    def load_data(self):
        if not os.path.exists(self.log_file):
            try:
                with file_lock:
                    with open(self.log_file, 'w', newline='', encoding='utf-8-sig') as f:
                        writer = csv.writer(f)
                        writer.writerow(['时间', '工作类别', '工作内容'])
            except Exception as e:
                CustomMessageBox(self, "错误", f"创建日志文件失败: {str(e)}").exec()

    def log_work(self, category):
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        content = ""
        
        if category == "其他":
            dialog = QDialog(self)
            dialog.setWindowTitle("输入工作内容")
            dialog.resize(400, 200)
            
            dialog_layout = QVBoxLayout(dialog)
            dialog_layout.addWidget(QLabel("请输入工作内容:"))
            
            text_edit = QTextEdit()
            dialog_layout.addWidget(text_edit)
            
            save_btn = QPushButton("保存")
            save_btn.clicked.connect(lambda: self.save_other_content(dialog, text_edit, current_time, category))
            dialog_layout.addWidget(save_btn)
            
            dialog.exec()
        else:
            self.save_log_entry(current_time, category, content)
    
    def save_other_content(self, dialog, text_edit, time, category):
        content = text_edit.toPlainText()
        if content.strip():
            self.save_log_entry(time, category, content)
            dialog.accept()
        else:
            CustomMessageBox(dialog, "警告", "工作内容不能为空").exec()
    
    def save_log_entry(self, time, category, content):
        try:
            with file_lock:
                if not os.path.exists(self.log_file):
                    self.load_data()
                
                with open(self.log_file, 'a', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow([time, category, content])
            
            CustomMessageBox(self, "成功", "工作日志已记录！").exec()
        except Exception as e:
            CustomMessageBox(self, "错误", f"保存日志失败: {str(e)}").exec()

    def undo_last_log(self):
        try:
            with file_lock:
                if not os.path.exists(self.log_file):
                    CustomMessageBox(self, "提示", "日志文件不存在").exec()
                    return

                lines = []
                with open(self.log_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    lines = list(reader)

                if len(lines) <= 1:
                    CustomMessageBox(self, "提示", "没有可撤销的记录").exec()
                    return

                last_row = lines[-1]
                if not last_row:
                    return

                log_time_str = last_row[0]
                try:
                    log_time = datetime.strptime(log_time_str, "%Y-%m-%d %H:%M:%S")
                    if (datetime.now() - log_time).total_seconds() <= 60:
                        lines.pop()
                        with open(self.log_file, 'w', newline='', encoding='utf-8-sig') as f:
                            writer = csv.writer(f)
                            writer.writerows(lines)
                        CustomMessageBox(self, "成功", "已撤销上一条记录").exec()
                    else:
                        CustomMessageBox(self, "失败", "只能撤销1分钟内的记录").exec()
                except ValueError:
                    CustomMessageBox(self, "错误", "无法解析最后一条记录的时间").exec()

        except Exception as e:
            CustomMessageBox(self, "错误", f"撤销失败: {str(e)}").exec()
    
    def export_stats_excel(self):
        try:
            if self.stats_table.rowCount() == 0:
                CustomMessageBox(self, "警告", "请先生成统计数据").exec()
                return
    
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存Excel文件",
                os.path.expanduser("~/Documents/WorkLog/工作统计.xlsx"),
                "Excel文件 (*.xlsx)"
            )
    
            if file_path:
                try:
                    from openpyxl import Workbook
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "工作统计"
                    
                    headers = ['工作类别', '次数', '归档统计']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    for row in range(self.stats_table.rowCount()):
                        category = self.stats_table.item(row, 0).text()
                        count = int(self.stats_table.item(row, 1).text())
                        archive = self.stats_table.item(row, 2).text()
                        
                        ws.cell(row=row+2, column=1, value=category)
                        ws.cell(row=row+2, column=2, value=count)
                        ws.cell(row=row+2, column=3, value=archive)
                    
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 10
                    ws.column_dimensions['C'].width = 25
                    
                    wb.save(file_path)
                    CustomMessageBox(self, "成功", f"统计数据已导出至 {file_path}").exec()
                    
                except ImportError:
                    CustomMessageBox(self, "错误", "导出Excel需要安装openpyxl库\n请运行: pip install openpyxl").exec()
                except Exception as e:
                    CustomMessageBox(self, "错误", f"导出Excel失败: {str(e)}").exec()
    
        except Exception as e:
            CustomMessageBox(self, "错误", f"导出Excel失败: {str(e)}").exec()

    def generate_stats(self):
        try:
            start = self.stats_start_date.date().toString("yyyy-MM-dd")
            end = self.stats_end_date.date().toString("yyyy-MM-dd")

            category_counts = {}
            try:
                with file_lock:
                    with open(self.log_file, 'r', encoding='utf-8-sig') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            try:
                                row_date = datetime.strptime(row['时间'], '%Y-%m-%d %H:%M:%S').date()
                                start_date = datetime.strptime(start, '%Y-%m-%d').date()
                                end_date = datetime.strptime(end, '%Y-%m-%d').date()
                                
                                if start_date <= row_date <= end_date:
                                    category = row['工作类别']
                                    category_counts[category] = category_counts.get(category, 0) + 1
                            except ValueError:
                                continue
            except Exception as e:
                CustomMessageBox(self, "错误", f"读取日志文件失败: {str(e)}").exec()
                return

            if len(category_counts) == 0:
                CustomMessageBox(self, "提示", "所选时间段内没有日志记录").exec()
                return

            for category in self.categories:
                if category not in category_counts:
                    category_counts[category] = 0

            self.stats_table.setRowCount(len(category_counts))
            for i, (category, count) in enumerate(category_counts.items()):
                archive_text = f"{category}{count}次"

                self.stats_table.setItem(i, 0, QTableWidgetItem(category))
                self.stats_table.setItem(i, 1, QTableWidgetItem(str(count)))
                self.stats_table.setItem(i, 2, QTableWidgetItem(archive_text))

            self.stats_table.resizeColumnsToContents()

        except Exception as e:
            CustomMessageBox(self, "错误", f"生成统计失败: {str(e)}").exec()

    def closeEvent(self, event):
        if self.server_thread and self.server_thread.isRunning():
            self.server_thread.stop()
            self.server_thread.wait()
        super().closeEvent(event)

    def toggle_auto_start(self, state):
        self.settings.setValue("auto_start_mobile_sync", bool(state))

    def check_auto_start(self):
        if self.settings.value("auto_start_mobile_sync", False, type=bool):
            # 60秒后自动开启
            QTimer.singleShot(60000, lambda: self.server_btn.setChecked(True))

class CustomMessageBox(QDialog):
    def __init__(self, parent=None, title="", message=""):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setFixedSize(300, 150)

        layout = QVBoxLayout(self)

        self.message_label = QLabel(message)
        self.message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.message_label)

        self.ok_button = QPushButton("确定")
        self.ok_button.clicked.connect(self.accept)
        layout.addWidget(self.ok_button)

        self.ok_button.setDefault(True)
        self.ok_button.setFocus()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Space or event.key() == Qt.Key.Key_Return:
            self.accept()
        else:
            super().keyPressEvent(event)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = WorkLogRecorder()
    window.show()
    sys.exit(app.exec())